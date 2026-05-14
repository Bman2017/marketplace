#!/usr/bin/env python3
"""Scrape the public UPenn Technology Publisher catalog into CSV, JSONL, and XLSX.

Public source: https://upenn.technologypublisher.com/
This reads public catalog/detail pages only. It does not bypass login or private systems.
"""
from __future__ import annotations

import argparse
import csv
import html
import json
import re
import sys
import time
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import quote, urljoin

import requests
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://upenn.technologypublisher.com/"
RSS_URL = "https://upenn.technologypublisher.com/rss.aspx"
TECH_URL_RE = re.compile(r"https?://upenn\.technologypublisher\.com/technology/\d+|/technology/\d+")
EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)
PHONE_RE = re.compile(r"(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}")
DATE_RE = re.compile(r"\b\d{1,2}/\d{1,2}/\d{4}\b|\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\s+\d{1,2},\s+\d{4}\b", re.I)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; ArnsInnovationsResearchBot/1.0; public university licensing catalog research)",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

FIELDS = [
    "source_university", "source_site", "technology_id", "url", "title", "short_description",
    "problem", "solution", "technology_overview", "advantages", "stage_of_development",
    "intellectual_property", "reference_media", "desired_partnerships", "case_id", "web_published",
    "docket", "contact_name", "contact_title", "contact_email", "contact_phone", "researchers",
    "categories", "patent_links", "source_rss_title", "source_rss_pubdate", "raw_section_text",
    "scrape_status", "error"
]

@dataclass
class Listing:
    source_university: str = "University of Pennsylvania"
    source_site: str = BASE_URL
    technology_id: str = ""
    url: str = ""
    title: str = ""
    short_description: str = ""
    problem: str = ""
    solution: str = ""
    technology_overview: str = ""
    advantages: str = ""
    stage_of_development: str = ""
    intellectual_property: str = ""
    reference_media: str = ""
    desired_partnerships: str = ""
    case_id: str = ""
    web_published: str = ""
    docket: str = ""
    contact_name: str = ""
    contact_title: str = ""
    contact_email: str = ""
    contact_phone: str = ""
    researchers: str = ""
    categories: str = ""
    patent_links: str = ""
    source_rss_title: str = ""
    source_rss_pubdate: str = ""
    raw_section_text: str = ""
    scrape_status: str = "ok"
    error: str = ""


def clean(s: str) -> str:
    s = html.unescape(s or "")
    s = s.replace("\xa0", " ")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def join_unique(vals: Iterable[str]) -> str:
    out = []
    for v in vals:
        v = clean(v)
        if v and v not in out:
            out.append(v)
    return " | ".join(out)


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def get(session: requests.Session, url: str, delay: float, timeout: int = 40) -> requests.Response:
    time.sleep(delay)
    r = session.get(url, timeout=timeout)
    r.raise_for_status()
    return r


def norm_url(url: str) -> str:
    return urljoin(BASE_URL, url.strip())


def sort_urls(urls: Iterable[str]) -> List[str]:
    def key(u: str) -> int:
        m = re.search(r"/technology/(\d+)", u)
        return int(m.group(1)) if m else 0
    return sorted(set(norm_url(u) for u in urls if "/technology/" in u), key=key)


def discover_from_rss(session: requests.Session, delay: float) -> Tuple[List[str], Dict[str, Dict[str, str]]]:
    urls: List[str] = []
    rss_meta: Dict[str, Dict[str, str]] = {}
    resp = get(session, RSS_URL, delay)
    text = resp.text
    # XML parse first.
    try:
        root = ET.fromstring(text.encode("utf-8"))
        for item in root.findall(".//item"):
            title = clean(item.findtext("title") or "")
            link = clean(item.findtext("link") or item.findtext("guid") or "")
            pub = clean(item.findtext("pubDate") or "")
            if link and "/technology/" in link:
                url = norm_url(link)
                urls.append(url)
                rss_meta[url] = {"title": title, "pubDate": pub}
    except Exception as e:
        print(f"RSS XML parse fallback: {e}", file=sys.stderr)
    # Regex fallback in case XML namespaces or malformed items appear.
    for m in TECH_URL_RE.finditer(text):
        urls.append(norm_url(m.group(0)))
    return sort_urls(urls), rss_meta


def script_urls(html_text: str) -> List[str]:
    soup = BeautifulSoup(html_text, "lxml")
    return [urljoin(BASE_URL, s.get("src")) for s in soup.find_all("script", src=True)]


def discover_algolia_config(session: requests.Session, delay: float) -> Dict[str, str]:
    blobs: List[str] = []
    home = get(session, BASE_URL, delay).text
    blobs.append(home)
    for js in script_urls(home):
        try:
            blobs.append(get(session, js, delay).text)
        except Exception as e:
            print(f"Skipping JS {js}: {e}", file=sys.stderr)
    combined = "\n".join(blobs)
    pats = {
        "app_id": [
            r"applicationID\s*[:=]\s*['\"]([A-Z0-9]{5,})['\"]",
            r"applicationId\s*[:=]\s*['\"]([A-Z0-9]{5,})['\"]",
            r"appId\s*[:=]\s*['\"]([A-Z0-9]{5,})['\"]",
            r"algoliasearch\(\s*['\"]([A-Z0-9]{5,})['\"]",
        ],
        "api_key": [
            r"searchApiKey\s*[:=]\s*['\"]([A-Za-z0-9_-]{20,})['\"]",
            r"searchAPIKey\s*[:=]\s*['\"]([A-Za-z0-9_-]{20,})['\"]",
            r"apiKey\s*[:=]\s*['\"]([A-Za-z0-9_-]{20,})['\"]",
            r"algoliasearch\(\s*['\"][A-Z0-9]{5,}['\"]\s*,\s*['\"]([A-Za-z0-9_-]{20,})['\"]",
        ],
        "index_name": [
            r"['\"](Test_Inteum_Tech_Publisher_PCI)['\"]",
            r"indexName\s*[:=]\s*['\"]([^'\"]*Tech_Publisher[^'\"]*)['\"]",
            r"index\s*[:=]\s*['\"]([^'\"]*Tech_Publisher[^'\"]*)['\"]",
        ],
    }
    out: Dict[str, str] = {}
    for k, patterns in pats.items():
        for pat in patterns:
            m = re.search(pat, combined, flags=re.I | re.S)
            if m:
                out[k] = m.group(1)
                break
    return out


def walk(obj: Any):
    if isinstance(obj, dict):
        for v in obj.values():
            yield from walk(v)
    elif isinstance(obj, list):
        for v in obj:
            yield from walk(v)
    elif isinstance(obj, str):
        yield obj


def hit_to_url(hit: Dict[str, Any]) -> Optional[str]:
    for key in ["url", "URL", "link", "href", "permalink", "direct_link", "technology_url", "path"]:
        val = hit.get(key)
        if isinstance(val, str) and "/technology/" in val:
            return norm_url(val)
    for val in walk(hit):
        m = TECH_URL_RE.search(val)
        if m:
            return norm_url(m.group(0))
    for key in ["objectID", "id", "technology_id", "technologyId", "record_id"]:
        val = hit.get(key)
        if val is not None and re.fullmatch(r"\d+", str(val)):
            return norm_url(f"/technology/{val}")
    return None


def discover_from_algolia(session: requests.Session, app_id: str, api_key: str, index_name: str, max_records: Optional[int]) -> List[str]:
    endpoint = f"https://{app_id.lower()}-dsn.algolia.net/1/indexes/{quote(index_name, safe='')}/query"
    headers = {
        "X-Algolia-Application-Id": app_id,
        "X-Algolia-API-Key": api_key,
        "Content-Type": "application/json",
        "User-Agent": HEADERS["User-Agent"],
    }
    urls: List[str] = []
    page = 0
    while True:
        payload = {"params": f"query=&hitsPerPage=100&page={page}&attributesToRetrieve=*"}
        r = session.post(endpoint, headers=headers, data=json.dumps(payload), timeout=40)
        r.raise_for_status()
        data = r.json()
        hits = data.get("hits", [])
        for h in hits:
            u = hit_to_url(h)
            if u:
                urls.append(u)
        print(f"Algolia page {page+1}/{data.get('nbPages')} hits={len(hits)} urls={len(set(urls))}", file=sys.stderr)
        if max_records and len(set(urls)) >= max_records:
            break
        page += 1
        if not hits or (data.get("nbPages") is not None and page >= int(data["nbPages"])):
            break
    return sort_urls(urls)[:max_records] if max_records else sort_urls(urls)


def discover_urls(session: requests.Session, delay: float, max_records: Optional[int]) -> Tuple[List[str], Dict[str, Any], Dict[str, Dict[str, str]]]:
    log: Dict[str, Any] = {"methods": []}
    rss_meta: Dict[str, Dict[str, str]] = {}
    urls: List[str] = []
    try:
        rss_urls, rss_meta = discover_from_rss(session, delay)
        log["methods"].append({"method": "rss", "count": len(rss_urls)})
        urls.extend(rss_urls)
    except Exception as e:
        log["methods"].append({"method": "rss", "error": repr(e)})
    try:
        cfg = discover_algolia_config(session, delay)
        log["algolia_config"] = {k: ("<redacted>" if k == "api_key" else v) for k, v in cfg.items()}
        if {"app_id", "api_key", "index_name"}.issubset(cfg):
            algolia_urls = discover_from_algolia(session, cfg["app_id"], cfg["api_key"], cfg["index_name"], max_records)
            log["methods"].append({"method": "algolia", "count": len(algolia_urls)})
            urls.extend(algolia_urls)
    except Exception as e:
        log["methods"].append({"method": "algolia", "error": repr(e)})
    urls = sort_urls(urls)
    if max_records:
        urls = urls[:max_records]
    log["discovered_url_count"] = len(urls)
    return urls, log, rss_meta


def heading_key(t: str) -> Optional[str]:
    k = clean(t).lower().strip(" :#")
    return {
        "problem": "problem",
        "solution": "solution",
        "technology": "technology_overview",
        "technology overview": "technology_overview",
        "advantages": "advantages",
        "stage of development": "stage_of_development",
        "intellectual property": "intellectual_property",
        "reference media": "reference_media",
        "desired partnerships": "desired_partnerships",
    }.get(k)


def extract_sections(soup: BeautifulSoup) -> Dict[str, str]:
    out: Dict[str, List[str]] = {}
    for h in soup.find_all(["h2", "h3", "h4"]):
        key = heading_key(h.get_text(" ", strip=True))
        if not key:
            continue
        parts: List[str] = []
        for sib in h.next_siblings:
            if isinstance(sib, Tag) and sib.name in ["h1", "h2", "h3", "h4"]:
                break
            text = sib.get_text(" ", strip=True) if isinstance(sib, Tag) else str(sib).strip()
            if clean(text):
                parts.append(clean(text))
        out[key] = parts
    return {k: clean("\n".join(v)) for k, v in out.items()}


def extract_between(lines: List[str], start: str, ends: List[str]) -> List[str]:
    start_l = start.lower()
    ends_l = [e.lower() for e in ends]
    start_i = None
    for i, line in enumerate(lines):
        if clean(line).lower() == start_l:
            start_i = i + 1
            break
    if start_i is None:
        return []
    out: List[str] = []
    for line in lines[start_i:]:
        if clean(line).lower() in ends_l:
            break
        if clean(line):
            out.append(clean(line))
    return out


def parse_detail(html_text: str, url: str, rss_meta: Optional[Dict[str, str]] = None) -> Listing:
    soup = BeautifulSoup(html_text, "lxml")
    rec = Listing(url=url)
    m = re.search(r"/technology/(\d+)", url)
    if m:
        rec.technology_id = m.group(1)
    if rss_meta:
        rec.source_rss_title = rss_meta.get("title", "")
        rec.source_rss_pubdate = rss_meta.get("pubDate", "")

    h1 = soup.find("h1")
    if h1:
        rec.title = clean(h1.get_text(" ", strip=True))
        desc = []
        for sib in h1.next_siblings:
            if isinstance(sib, Tag) and sib.name in ["h2", "h3", "h4"]:
                break
            txt = sib.get_text(" ", strip=True) if isinstance(sib, Tag) else str(sib).strip()
            txt = clean(txt)
            if txt and "BACK TO SEARCH" not in txt and "Download as PDF" not in txt:
                desc.append(txt)
        rec.short_description = clean(" ".join(desc))

    for k, v in extract_sections(soup).items():
        setattr(rec, k, v)

    all_text = soup.get_text("\n", strip=True)
    rec.raw_section_text = clean(all_text)
    lines = [clean(x) for x in all_text.split("\n") if clean(x)]
    flat = "\n".join(lines)

    for attr, pat in [
        ("case_id", r"Case ID:\s*([^\n]+)"),
        ("web_published", r"Web Published:\s*([^\n]+)"),
        ("docket", r"Docket(?:\s*(?:No\.|#|:))?\s*[:#]?\s*([^\n]+)"),
    ]:
        mm = re.search(pat, flat, flags=re.I)
        if mm:
            setattr(rec, attr, clean(mm.group(1)))
    if not rec.web_published:
        dm = DATE_RE.search(flat)
        if dm:
            rec.web_published = dm.group(0)

    contact_lines = extract_between(lines, "Contact", ["RESEARCHERS", "CATEGORY(S)", "Keywords", "Docket"])
    emails = [EMAIL_RE.search(x).group(0) for x in contact_lines if EMAIL_RE.search(x)]
    phones = [PHONE_RE.search(x).group(0) for x in contact_lines if PHONE_RE.search(x)]
    useful = [x for x in contact_lines if not EMAIL_RE.search(x) and not PHONE_RE.search(x) and x != "University of Pennsylvania"]
    if useful:
        rec.contact_name = useful[0]
    if len(useful) > 1:
        rec.contact_title = useful[1]
    if emails:
        rec.contact_email = emails[0]
    if phones:
        rec.contact_phone = phones[0]

    rec.researchers = join_unique(extract_between(lines, "RESEARCHERS", ["CATEGORY(S)", "Keywords", "Docket", "Contact"]))
    cats = []
    for a in soup.find_all("a"):
        t = clean(a.get_text(" ", strip=True))
        if "Technology Classifications >" in t:
            cats.append(t.replace("Technology Classifications >", "").strip())
    if not cats:
        cats = [x.replace("Technology Classifications >", "").strip() for x in extract_between(lines, "CATEGORY(S)", ["Keywords", "Docket", "Contact"])]
    rec.categories = join_unique(cats)

    patents = []
    for a in soup.find_all("a", href=True):
        href = a.get("href", "")
        label = clean(a.get_text(" ", strip=True))
        if "patents.google.com" in href or re.search(r"\b(?:US|WO|EP)\s*\d", label, flags=re.I):
            patents.append(f"{label} ({href})" if label else href)
    rec.patent_links = join_unique(patents)
    return rec


def write_csv(records: List[Listing], output_dir: Path) -> Path:
    path = output_dir / "upenn_technology_listings.csv"
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS, extrasaction="ignore")
        w.writeheader()
        for r in records:
            w.writerow(asdict(r))
    return path


def write_jsonl(records: List[Listing], output_dir: Path) -> Path:
    path = output_dir / "upenn_technology_listings.jsonl"
    with path.open("w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(asdict(r), ensure_ascii=False) + "\n")
    return path


def write_xlsx(records: List[Listing], output_dir: Path, log: Dict[str, Any]) -> Path:
    path = output_dir / "upenn_technology_listings.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "UPenn Listings"
    ws.append(FIELDS)
    for r in records:
        d = asdict(r)
        ws.append([d.get(f, "") for f in FIELDS])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 24, "B": 34, "C": 14, "D": 56, "E": 56, "F": 70, "G": 75, "H": 75,
        "I": 75, "J": 60, "K": 28, "L": 45, "M": 45, "N": 40, "O": 22, "P": 20,
        "Q": 22, "R": 28, "S": 36, "T": 34, "U": 20, "V": 48, "W": 55, "X": 70,
        "Y": 45, "Z": 28, "AA": 90, "AB": 18, "AC": 48,
    }
    for col_idx in range(1, len(FIELDS) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(letter, 25)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[1].height = 30

    summary = wb.create_sheet("Summary")
    summary_rows = [
        ("Source", BASE_URL),
        ("RSS", RSS_URL),
        ("Discovered URLs", log.get("discovered_url_count", "")),
        ("Records written", len(records)),
        ("OK records", sum(1 for r in records if r.scrape_status == "ok")),
        ("Error records", sum(1 for r in records if r.scrape_status != "ok")),
        ("Generated by", "GitHub Actions scraper in Bman2017/marketplace"),
    ]
    summary.append(["Metric", "Value"])
    for row in summary_rows:
        summary.append(list(row))
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 90

    wb.save(path)
    return path


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--output-dir", default="data/upenn")
    p.add_argument("--delay", type=float, default=0.35)
    p.add_argument("--max", type=int, default=None)
    args = p.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    session = make_session()
    log: Dict[str, Any] = {"base_url": BASE_URL, "rss_url": RSS_URL, "started_at_unix": time.time(), "errors": []}

    urls, discover_log, rss_meta = discover_urls(session, args.delay, args.max)
    log.update(discover_log)
    if not urls:
        raise RuntimeError("No UPenn technology URLs discovered from RSS or public search index.")

    records: List[Listing] = []
    for i, url in enumerate(urls, 1):
        print(f"[{i}/{len(urls)}] {url}", file=sys.stderr)
        try:
            resp = get(session, url, args.delay)
            rec = parse_detail(resp.text, url, rss_meta.get(url, {}))
            records.append(rec)
        except Exception as e:
            rec = Listing(url=url)
            m = re.search(r"/technology/(\d+)", url)
            if m:
                rec.technology_id = m.group(1)
            rec.scrape_status = "error"
            rec.error = repr(e)
            records.append(rec)
            log["errors"].append({"url": url, "error": repr(e)})

    with (output_dir / "discovered_urls.txt").open("w", encoding="utf-8") as f:
        for u in urls:
            f.write(u + "\n")
    log["record_count"] = len(records)
    log["ok_count"] = sum(1 for r in records if r.scrape_status == "ok")
    log["error_count"] = sum(1 for r in records if r.scrape_status != "ok")
    log["finished_at_unix"] = time.time()
    write_csv(records, output_dir)
    write_jsonl(records, output_dir)
    write_xlsx(records, output_dir, log)
    with (output_dir / "scrape_log.json").open("w", encoding="utf-8") as f:
        json.dump(log, f, indent=2, ensure_ascii=False)
    print(json.dumps({"records": len(records), "ok": log["ok_count"], "errors": log["error_count"]}, indent=2))


if __name__ == "__main__":
    main()
